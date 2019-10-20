import sys
import os
import pandas as pd
import xlsxwriter
import openpyxl
sys.path.append(os.path.abspath('./balancing_dataset'))
import absenceData

import unittest

'''
This test is used to test whether there is a balanced amount of pseudo absence data compared to VBA
To run this test, run
    python absence_data_test.py
'''
class TestPseudoAbsenceDataLength(unittest.TestCase):

    '''
    Test for Agile Antechinus
    '''
    def test_absencedata_generation_length_agile_antechinus(self):
        absenceData.generateAbsenceData('Agile Antechinus', './tests/testing_absence/agile_antechinus_test.xlsx')
        wb = openpyxl.load_workbook('./tests/testing_absence/agile_antechinus_test.xlsx')
        existingWorksheet = wb['Sheet1']
        row_count = existingWorksheet.max_row

        vba = pd.read_excel('./VBA_Raster.xlsx')
        ufi = vba.loc[vba['COMMON_NME'] == "Agile Antechinus"]
        numOfItem = len(ufi)

        self.assertGreaterEqual(numOfItem, row_count)

    '''
    Test for Brown_treecreeper
    '''
    def test_absencedata_generation_length_brown_treecreeper(self):
        absenceData.generateAbsenceData('Brown Treecreeper', './tests/testing_absence/brown_treecreeper_test.xlsx');
        wb = openpyxl.load_workbook('./tests/testing_absence/brown_treecreeper_test.xlsx')
        existingWorksheet = wb['Sheet1']
        row_count = existingWorksheet.max_row

        vba = pd.read_excel('./VBA_Raster.xlsx')
        ufi = vba.loc[vba['COMMON_NME'] == "Brown Treecreeper"]
        numOfItem = len(ufi)

        self.assertGreaterEqual(numOfItem, row_count)

    
    '''
    Test for Common Beard heath
    '''
    def test_absencedata_generation_length_common_beard_heath(self):
        absenceData.generateAbsenceData('Common Beard-heath', './tests/testing_absence/common_beard_heath_test.xlsx');
        wb = openpyxl.load_workbook('./tests/testing_absence/common_beard_heath_test.xlsx')
        existingWorksheet = wb['Sheet1']
        row_count = existingWorksheet.max_row

        vba = pd.read_excel('./VBA_Raster.xlsx')
        ufi = vba.loc[vba['COMMON_NME'] == "Common Beard-heath"]
        numOfItem = len(ufi)

        self.assertGreaterEqual(row_count, numOfItem)

        
    '''
    Test for small triggerplant 
    '''
    def test_absencedata_generation_length_small_triggerplant_test(self):
        absenceData.generateAbsenceData('Small Triggerplant', './tests/testing_absence/small_triggerplant_test.xlsx');
        wb = openpyxl.load_workbook('./tests/testing_absence/small_triggerplant_test.xlsx')
        existingWorksheet = wb['Sheet1']
        row_count = existingWorksheet.max_row

        vba = pd.read_excel('./VBA_Raster.xlsx')
        ufi = vba.loc[vba['COMMON_NME'] == "Small Triggerplant"]
        numOfItem = len(ufi)

        self.assertGreaterEqual(row_count, numOfItem)

    
    '''
    Test for tree frog
    '''
    def test_absencedata_generation_length_tree_frog_test(self):
        absenceData.generateAbsenceData('Southern Brown Tree Frog', './tests/testing_absence/tree_frog_test.xlsx');
        wb = openpyxl.load_workbook('./tests/testing_absence/tree_frog_test.xlsx')
        existingWorksheet = wb['Sheet1']
        row_count = existingWorksheet.max_row

        vba = pd.read_excel('./VBA_Raster.xlsx')
        ufi = vba.loc[vba['COMMON_NME'] == "Southern Brown Tree Frog"]
        numOfItem = len(ufi)

        self.assertGreaterEqual(row_count, numOfItem)

    '''
    Test for white browed treecreeper
    '''
    def test_absencedata_generation_length_white_browed_treecreeper_test(self):
        absenceData.generateAbsenceData('White-browed Treecreeper', './tests/testing_absence/white_browed_treecreeper_test.xlsx');
        wb = openpyxl.load_workbook('./tests/testing_absence/white_browed_treecreeper_test.xlsx')
        existingWorksheet = wb['Sheet1']
        row_count = existingWorksheet.max_row

        vba = pd.read_excel('./VBA_Raster.xlsx')
        ufi = vba.loc[vba['COMMON_NME'] == "White-browed Treecreeper"]
        numOfItem = len(ufi)

        self.assertGreaterEqual(row_count, numOfItem)

        

if __name__ == '__main__':
    unittest.main()
    